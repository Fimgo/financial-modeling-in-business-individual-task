from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_FILE = BASE_DIR / "Starbucks_Financial_Model.xlsx"


BLUE = "1F4E78"
DARK = "17365D"
LIGHT_BLUE = "D9EAF7"
LIGHT_GREEN = "E2F0D9"
LIGHT_YELLOW = "FFF2CC"
LIGHT_RED = "FCE4D6"
WHITE = "FFFFFF"
BLACK = "000000"
INPUT_BLUE = "0000FF"
LINK_GREEN = "008000"


def style_title(cell):
    cell.font = Font(bold=True, size=16, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_section(cell):
    cell.font = Font(bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=DARK)
    cell.alignment = Alignment(horizontal="left")


def style_header(cell):
    cell.font = Font(bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_input(cell):
    cell.font = Font(color=INPUT_BLUE)
    cell.fill = PatternFill("solid", fgColor=LIGHT_YELLOW)


def style_link(cell):
    cell.font = Font(color=LINK_GREEN)


def set_common_sheet_style(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B4"
    thin_gray = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(bottom=thin_gray)
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions["A"].width = 36


def write_title(ws, title, subtitle=None):
    ws.merge_cells("A1:H1")
    ws["A1"] = title
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 26
    if subtitle:
        ws.merge_cells("A2:H2")
        ws["A2"] = subtitle
        ws["A2"].font = Font(italic=True, color="666666")


historical = {
    2023: {
        "Revenue": 35975.6,
        "Product and distribution costs": 11409.1,
        "Store operating expenses": 14720.3,
        "Other operating expenses": 539.4,
        "Depreciation and amortization": 1362.6,
        "General and administrative expenses": 2441.3,
        "Restructuring and impairments": 21.8,
        "Income from equity investees": 298.4,
    },
    2024: {
        "Revenue": 36176.2,
        "Product and distribution costs": 11180.6,
        "Store operating expenses": 15286.5,
        "Other operating expenses": 565.6,
        "Depreciation and amortization": 1512.6,
        "General and administrative expenses": 2523.3,
        "Restructuring and impairments": 0.0,
        "Income from equity investees": 301.2,
    },
    2025: {
        "Revenue": 37184.4,
        "Product and distribution costs": 11658.2,
        "Store operating expenses": 17058.9,
        "Other operating expenses": 584.6,
        "Depreciation and amortization": 1684.7,
        "General and administrative expenses": 2617.2,
        "Restructuring and impairments": 892.0,
        "Income from equity investees": 247.8,
    },
}

source_rows = [
    [
        "S1",
        "Starbucks Fiscal 2025 Annual Report PDF",
        "https://s203.q4cdn.com/326826266/files/doc_financials/2025/ar/Starbucks-Corporation_2025-Annual-Report-Web-Ready.pdf",
        "Consolidated revenues, operating expenses, operating income, net earnings, segment and business description.",
    ],
    [
        "S2",
        "Starbucks Investor Relations - Annual Reports",
        "https://investor.starbucks.com/financials/annual-reports/",
        "Official page for Starbucks annual reports.",
    ],
    [
        "S3",
        "Starbucks SEC Filing Details - Form 10-K",
        "https://investor.starbucks.com/financials/sec-filings/sec-filings-details/default.aspx?FilingId=18927461",
        "Official 10-K filing page dated 2025-11-14 with PDF, Excel, and XBRL downloads.",
    ],
]


wb = Workbook()
wb.remove(wb.active)
for name in [
    "Cover",
    "Data",
    "Assumptions",
    "P&L",
    "Break-even",
    "Sensitivity",
    "Charts",
    "Checks",
    "Sources",
]:
    wb.create_sheet(name)

wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True

# Cover
ws = wb["Cover"]
write_title(ws, "Starbucks Corporation Financial Model", "Навчальна базова фінансова модель за даними реального підприємства")
cover_items = [
    ("Компанія", "Starbucks Corporation"),
    ("Тікер", "SBUX"),
    ("Галузь", "Ресторани / спеціалізована кавова роздрібна торгівля"),
    ("Валюта та одиниці", "USD millions"),
    ("Історичний період", "FY2023-FY2025"),
    ("Прогнозний період", "FY2026E-FY2028E"),
    ("Основне джерело", "Starbucks Fiscal 2025 Annual Report, Form 10-K"),
    ("Статус моделі", "=Checks!F10"),
]
for i, (label, value) in enumerate(cover_items, start=4):
    ws.cell(i, 1, label)
    ws.cell(i, 2, value)
    ws.cell(i, 1).font = Font(bold=True)
ws["A14"] = "Коротка логіка"
style_section(ws["A14"])
ws.merge_cells("A14:H14")
ws["A15"] = (
    "Модель бере фактичні фінансові дані Starbucks за 2023-2025 роки, "
    "додає припущення для 2026-2028 років, формує спрощений P&L, "
    "рахує break-even revenue та показує чутливість прибутку до Revenue growth і Variable cost ratio."
)
ws.merge_cells("A15:H17")
ws["A15"].alignment = Alignment(wrap_text=True, vertical="top")
set_common_sheet_style(ws)

# Data
ws = wb["Data"]
write_title(ws, "Historical Data", "Actuals from Starbucks Fiscal 2025 Annual Report, USD millions")
headers = ["Metric", "2023A", "2024A", "2025A", "Source ID", "Notes"]
for col, header in enumerate(headers, start=1):
    ws.cell(3, col, header)
    style_header(ws.cell(3, col))
metrics = [
    "Revenue",
    "Product and distribution costs",
    "Store operating expenses",
    "Other operating expenses",
    "Depreciation and amortization",
    "General and administrative expenses",
    "Restructuring and impairments",
    "Total operating expenses",
    "Income from equity investees",
    "Operating income",
    "Operating margin",
    "Net earnings attributable to Starbucks",
]
net_income = {2023: 4124.5, 2024: 3760.9, 2025: 1856.4}
for row_idx, metric in enumerate(metrics, start=4):
    ws.cell(row_idx, 1, metric)
    for col_idx, year in enumerate([2023, 2024, 2025], start=2):
        if metric == "Total operating expenses":
            ws.cell(row_idx, col_idx, f"=SUM({get_column_letter(col_idx)}5:{get_column_letter(col_idx)}10)")
        elif metric == "Operating income":
            ws.cell(row_idx, col_idx, f"={get_column_letter(col_idx)}4-{get_column_letter(col_idx)}11+{get_column_letter(col_idx)}12")
        elif metric == "Operating margin":
            ws.cell(row_idx, col_idx, f"={get_column_letter(col_idx)}13/{get_column_letter(col_idx)}4")
        elif metric == "Net earnings attributable to Starbucks":
            ws.cell(row_idx, col_idx, net_income[year])
        else:
            ws.cell(row_idx, col_idx, historical[year][metric])
    ws.cell(row_idx, 5, "S1")
    ws.cell(row_idx, 6, "Form 10-K consolidated results of operations")
for row in range(4, 16):
    for col in range(2, 5):
        ws.cell(row, col).number_format = "0.0"
for col in range(2, 5):
    ws.cell(14, col).number_format = "0.0%"
set_common_sheet_style(ws)
ws.column_dimensions["F"].width = 42

# Assumptions
ws = wb["Assumptions"]
write_title(ws, "Assumptions", "Blue/yellow cells are editable assumptions")
headers = ["Driver", "2026E", "2027E", "2028E", "Rationale"]
for col, header in enumerate(headers, start=1):
    ws.cell(3, col, header)
    style_header(ws.cell(3, col))
assumption_rows = [
    ("Revenue growth", 0.035, 0.040, 0.040, "Moderate recovery after FY2025; conservative normalized growth."),
    ("Variable cost ratio", 0.765, 0.760, 0.755, "Costs tied to revenue: product, distribution, store, and other operating costs."),
    ("Product/distribution share of variable costs", 0.385, 0.385, 0.385, "Approximate split based on FY2025 cost structure."),
    ("Store operating share of variable costs", 0.595, 0.595, 0.595, "Main variable/semi-variable operating cost block."),
    ("Other operating share of variable costs", 0.020, 0.020, 0.020, "Small remaining operating cost category."),
    ("Fixed cost growth", 0.030, 0.030, 0.030, "D&A and G&A grow below revenue growth."),
    ("Equity income as % of revenue", 0.006, 0.006, 0.006, "Based on recent historical range."),
    ("Restructuring costs", 300.0, 100.0, 50.0, "Assumed decline after elevated FY2025 restructuring costs."),
]
for row_idx, row in enumerate(assumption_rows, start=4):
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row_idx, col_idx, value)
    for col in range(2, 5):
        style_input(ws.cell(row_idx, col))
        ws.cell(row_idx, col).number_format = "0.0%" if row_idx != 11 else "0.0"
ws["A14"] = "Convention"
style_section(ws["A14"])
ws.merge_cells("A14:E14")
ws["A15"] = "Forecast formulas reference these assumptions. Change blue/yellow cells to test another scenario."
ws.merge_cells("A15:E16")
ws["A15"].alignment = Alignment(wrap_text=True, vertical="top")
set_common_sheet_style(ws)
ws.column_dimensions["E"].width = 58

# P&L
ws = wb["P&L"]
write_title(ws, "Simplified P&L", "Historical actuals and forecast, USD millions")
headers = ["Metric", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E"]
for col, header in enumerate(headers, start=1):
    ws.cell(3, col, header)
    style_header(ws.cell(3, col))
pnl_rows = [
    "Revenue",
    "Product and distribution costs",
    "Store operating expenses",
    "Other operating expenses",
    "Depreciation and amortization",
    "General and administrative expenses",
    "Restructuring and impairments",
    "Total operating expenses",
    "Income from equity investees",
    "Operating income",
    "Operating margin",
]
for row_idx, metric in enumerate(pnl_rows, start=4):
    ws.cell(row_idx, 1, metric)
    for col_idx in range(2, 5):
        data_row = metrics.index(metric) + 4
        ws.cell(row_idx, col_idx, f"=Data!{get_column_letter(col_idx)}{data_row}")
        style_link(ws.cell(row_idx, col_idx))

for col_idx in range(5, 8):
    prev_col = get_column_letter(col_idx - 1)
    curr_col = get_column_letter(col_idx)
    ass_col = get_column_letter(col_idx - 3)
    ws.cell(4, col_idx, f"={prev_col}4*(1+Assumptions!{ass_col}$4)")
    ws.cell(5, col_idx, f"={curr_col}4*Assumptions!{ass_col}$5*Assumptions!{ass_col}$6")
    ws.cell(6, col_idx, f"={curr_col}4*Assumptions!{ass_col}$5*Assumptions!{ass_col}$7")
    ws.cell(7, col_idx, f"={curr_col}4*Assumptions!{ass_col}$5*Assumptions!{ass_col}$8")
    ws.cell(8, col_idx, f"={prev_col}8*(1+Assumptions!{ass_col}$9)")
    ws.cell(9, col_idx, f"={prev_col}9*(1+Assumptions!{ass_col}$9)")
    ws.cell(10, col_idx, f"=Assumptions!{ass_col}$11")
    ws.cell(11, col_idx, f"=SUM({curr_col}5:{curr_col}10)")
    ws.cell(12, col_idx, f"={curr_col}4*Assumptions!{ass_col}$10")
    ws.cell(13, col_idx, f"={curr_col}4-{curr_col}11+{curr_col}12")
    ws.cell(14, col_idx, f"={curr_col}13/{curr_col}4")
for row in range(4, 15):
    for col in range(2, 8):
        ws.cell(row, col).number_format = "0.0"
for col in range(2, 8):
    ws.cell(14, col).number_format = "0.0%"
for row in [11, 13]:
    for col in range(1, 8):
        ws.cell(row, col).font = Font(bold=True)
        ws.cell(row, col).fill = PatternFill("solid", fgColor=LIGHT_BLUE if row == 11 else LIGHT_GREEN)
set_common_sheet_style(ws)

# Break-even
ws = wb["Break-even"]
write_title(ws, "Break-even Analysis", "Break-even revenue based on contribution margin method")
items = [
    ("Metric", "Formula / Value", "Explanation"),
    ("Revenue 2026E", "='P&L'!E4", "Base forecast revenue"),
    ("Variable cost ratio 2026E", "=Assumptions!B5", "Variable costs as % of revenue"),
    ("Contribution margin %", "=1-B5", "1 - variable cost ratio"),
    ("Fixed costs 2026E", "='P&L'!E8+'P&L'!E9+'P&L'!E10", "D&A + G&A + restructuring"),
    ("Break-even revenue 2026E", "=B7/B6", "Fixed costs / contribution margin"),
    ("Safety margin vs forecast", "=B4-B8", "Revenue forecast minus break-even revenue"),
]
for row_idx, row in enumerate(items, start=3):
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row_idx, col_idx, value)
        if row_idx == 3:
            style_header(ws.cell(row_idx, col_idx))
for cell in ["B4", "B7", "B8", "B9"]:
    ws[cell].number_format = "0.0"
for cell in ["B5", "B6"]:
    ws[cell].number_format = "0.0%"
set_common_sheet_style(ws)
ws.column_dimensions["C"].width = 48

# Sensitivity
ws = wb["Sensitivity"]
write_title(ws, "Sensitivity Analysis", "2026E operating income sensitivity to revenue growth and variable cost ratio")
ws["A3"] = "Operating income 2026E, USD millions"
style_section(ws["A3"])
ws.merge_cells("A3:E3")
ws["A5"] = "Revenue growth delta"
ws["B4"] = "Variable cost ratio"
vc_ratios = [0.745, 0.765, 0.785]
growth_deltas = [-0.05, 0.00, 0.05]
for col_idx, ratio in enumerate(vc_ratios, start=2):
    ws.cell(5, col_idx, ratio)
    ws.cell(5, col_idx).number_format = "0.0%"
    style_header(ws.cell(5, col_idx))
for row_idx, delta in enumerate(growth_deltas, start=6):
    ws.cell(row_idx, 1, delta)
    ws.cell(row_idx, 1).number_format = "0.0%"
    style_header(ws.cell(row_idx, 1))
    for col_idx, ratio in enumerate(vc_ratios, start=2):
        col_letter = get_column_letter(col_idx)
        ws.cell(
            row_idx,
            col_idx,
            f"=('P&L'!D4*(1+Assumptions!B4+$A{row_idx}))*(1-{col_letter}$5)-('P&L'!E8+'P&L'!E9+Assumptions!B11)+('P&L'!D4*(1+Assumptions!B4+$A{row_idx}))*Assumptions!B10",
        )
        ws.cell(row_idx, col_idx).number_format = "0.0"
ws["A11"] = "Interpretation"
style_section(ws["A11"])
ws.merge_cells("A11:E11")
ws["A12"] = "Higher revenue growth improves operating income, while a higher variable cost ratio quickly compresses profitability."
ws.merge_cells("A12:E13")
ws["A12"].alignment = Alignment(wrap_text=True, vertical="top")
set_common_sheet_style(ws)

# Checks
ws = wb["Checks"]
write_title(ws, "Model Checks", "Basic source, formula, and logic checks")
headers = ["Check", "Actual", "Expected", "Difference", "Tolerance", "Status"]
for col, header in enumerate(headers, start=1):
    ws.cell(3, col, header)
    style_header(ws.cell(3, col))
checks = [
    ("2025 revenue ties to source", "='P&L'!D4", "=Data!D4", "=B4-C4", 0.01, '=IF(ABS(D4)<=E4,"OK","Review")'),
    ("2025 operating income ties to source formula", "='P&L'!D13", "=Data!D13", "=B5-C5", 0.01, '=IF(ABS(D5)<=E5,"OK","Review")'),
    ("2026 total opex sums components", "='P&L'!E11", "=SUM('P&L'!E5:E10)", "=B6-C6", 0.01, '=IF(ABS(D6)<=E6,"OK","Review")'),
    ("2026 operating margin formula", "='P&L'!E14", "='P&L'!E13/'P&L'!E4", "=B7-C7", 0.0001, '=IF(ABS(D7)<=E7,"OK","Review")'),
    ("Break-even safety margin formula", "='Break-even'!B9", "='Break-even'!B4-'Break-even'!B8", "=B8-C8", 0.01, '=IF(ABS(D8)<=E8,"OK","Review")'),
]
for row_idx, row in enumerate(checks, start=4):
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row_idx, col_idx, value)
ws["A10"] = "Overall model status"
ws["F10"] = '=IF(COUNTIF(F4:F8,"Review")=0,"OK","Review")'
ws["F10"].font = Font(bold=True, color=WHITE)
ws["F10"].fill = PatternFill("solid", fgColor="70AD47")
for row in range(4, 9):
    for col in range(2, 6):
        ws.cell(row, col).number_format = "0.000"
set_common_sheet_style(ws)

# Sources
ws = wb["Sources"]
write_title(ws, "Sources and Audit Trail", "Official Starbucks / SEC source links")
headers = ["Source ID", "Source", "URL", "Notes"]
for col, header in enumerate(headers, start=1):
    ws.cell(3, col, header)
    style_header(ws.cell(3, col))
for row_idx, row in enumerate(source_rows, start=4):
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row_idx, col_idx, value)
    ws.cell(row_idx, 3).hyperlink = row[2]
    ws.cell(row_idx, 3).style = "Hyperlink"
set_common_sheet_style(ws)
ws.column_dimensions["B"].width = 42
ws.column_dimensions["C"].width = 80
ws.column_dimensions["D"].width = 62
for row in range(4, 7):
    ws.row_dimensions[row].height = 42
    for col in range(1, 5):
        ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")

# Charts
ws = wb["Charts"]
write_title(ws, "Charts", "Revenue, costs, operating income, and margin")
ws["A3"] = "Charts are linked to the P&L sheet and update when assumptions change."
ws.merge_cells("A3:H3")

line = LineChart()
line.title = "Revenue Dynamics"
line.y_axis.title = "USD millions"
line.x_axis.title = "Fiscal year"
data = Reference(wb["P&L"], min_col=2, max_col=7, min_row=4, max_row=4)
cats = Reference(wb["P&L"], min_col=2, max_col=7, min_row=3, max_row=3)
line.add_data(data, from_rows=True, titles_from_data=False)
line.set_categories(cats)
line.height = 7
line.width = 16
ws.add_chart(line, "A5")

bar = BarChart()
bar.title = "Operating Income"
bar.y_axis.title = "USD millions"
data = Reference(wb["P&L"], min_col=2, max_col=7, min_row=13, max_row=13)
bar.add_data(data, from_rows=True, titles_from_data=False)
bar.set_categories(cats)
bar.height = 7
bar.width = 16
bar.dataLabels = DataLabelList()
bar.dataLabels.showVal = False
ws.add_chart(bar, "J5")

combo = LineChart()
combo.title = "Operating Margin"
combo.y_axis.title = "%"
data = Reference(wb["P&L"], min_col=2, max_col=7, min_row=14, max_row=14)
combo.add_data(data, from_rows=True, titles_from_data=False)
combo.set_categories(cats)
combo.height = 7
combo.width = 16
ws.add_chart(combo, "A22")

cost = BarChart()
cost.title = "Revenue vs Total Operating Expenses"
cost.y_axis.title = "USD millions"
data = Reference(wb["P&L"], min_col=2, max_col=7, min_row=4, max_row=11)
cost.add_data(data, from_rows=True, titles_from_data=True)
cost.set_categories(cats)
cost.height = 7
cost.width = 16
ws.add_chart(cost, "J22")
set_common_sheet_style(ws)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if sheet_name == "Assumptions" and cell.row in range(4, 11):
                    cell.number_format = "0.0%"
                elif sheet_name == "P&L" and cell.row == 14:
                    cell.number_format = "0.0%"
                elif sheet_name == "Data" and cell.row == 14:
                    cell.number_format = "0.0%"
                else:
                    cell.number_format = "0.0"

wb.save(OUTPUT_FILE)

# Lightweight verification of formulas and structure.
check_wb = load_workbook(OUTPUT_FILE, data_only=False)
assert set(["Cover", "Data", "Assumptions", "P&L", "Break-even", "Sensitivity", "Charts", "Checks", "Sources"]).issubset(check_wb.sheetnames)
assert check_wb["Data"]["D4"].value == 37184.4
assert check_wb["P&L"]["E4"].value == "=D4*(1+Assumptions!B$4)"
assert check_wb["Break-even"]["B8"].value == "=B7/B6"
assert check_wb["Checks"]["F10"].value == '=IF(COUNTIF(F4:F8,"Review")=0,"OK","Review")'

print("Created Starbucks_Financial_Model.xlsx")
